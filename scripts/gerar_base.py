from pathlib import Path
from datetime import datetime, timezone
from collections import Counter, defaultdict
from html import unescape
from urllib.parse import urljoin
import csv
import json
import os
import re
import time
import urllib.error
import urllib.request
import tempfile
import shutil
import ssl

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
DADOS = BASE / "dados"

LIMITE_REDUCAO = 0.20
METADADOS_FONTES = {}
VERSAO_GERADOR = "2026-08-30-cmed-v2"

URL_DISPOSITIVOS = (
    "https://dados.anvisa.gov.br/dados/"
    "TA_PRODUTO_SAUDE_SITE.csv"
)

URL_AFE_AE = (
    "https://dados.anvisa.gov.br/dados/"
    "CONSULTAS/EMPRESA_FISCALIZACAO_PRODUTO/"
    "TA_CONSULTA_FUNCIONAMENTO_EMPRESA_NACIONAL.CSV"
)

URL_MEDICAMENTOS = (
    "https://dados.anvisa.gov.br/dados/"
    "DADOS_ABERTOS_MEDICAMENTOS.csv"
)

URL_SANEANTES = (
    "https://dados.anvisa.gov.br/dados/"
    "CONSULTAS/PRODUTOS/"
    "TA_CONSULTA_SANEANTES.CSV"
)

URL_COSMETICOS = (
    "https://dados.anvisa.gov.br/dados/"
    "CONSULTAS/PRODUTOS/"
    "TA_CONSULTA_COSMETICOS.CSV"
)

URL_ALIMENTOS = (
    "https://dados.anvisa.gov.br/dados/"
    "CONSULTAS/PRODUTOS/"
    "TA_CONSULTA_ALIMENTOS.CSV"
)

URL_ALIMENTOS_RESULTADO = (
    "https://dados.anvisa.gov.br/dados/"
    "CONSULTAS/PRODUTOS/"
    "TA_CONSULTA_ALIMENTOS_RESULTADO.CSV"
)

PAGINA_CMED = (
    "https://www.gov.br/anvisa/pt-br/assuntos/"
    "medicamentos/cmed/precos"
)

# O prefixo é publicado no manifesto. Assim, o HTML funciona com a base atual
# de três dígitos e também após a redistribuição em fragmentos menores.
PREFIXOS_BASE = {
    "dispositivos": 5,
    "afe_ae": 3,
    "medicamentos": 4,
    "saneantes": 4,
    # Cosméticos são fragmentados pelos dígitos 6 a 8 do processo,
    # e não pelos primeiros dígitos (que quase sempre são 25351).
    "cosmeticos": 3,
    # Quatro dígitos evitam concentrar processos antigos em arquivos grandes.
    "alimentos": 4,
    # O GTIN é normalizado com 14 dígitos e consultado diretamente.
    "cmed": 8,
}
PREFIXO_INDICE_CNPJ = 3
PREFIXO_INDICE_AUTORIZACAO = 3
PREFIXO_INDICE_REGULARIZACAO_ALIMENTOS = 3
TETO_FRAGMENTO = 300_000
def somente_numeros(valor):
    return re.sub(r"\D", "", str(valor or ""))


def texto(valor):
    return re.sub(
        r"\s+",
        " ",
        str(valor or "").strip()
    )


def normalizar_nome(valor):
    return re.sub(
        r"[^A-Z0-9]",
        "",
        str(valor or "").upper()
    )


def achar_coluna(cabecalho, possibilidades):
    mapa = {
        normalizar_nome(c): c
        for c in cabecalho
        if c
    }

    for nome in possibilidades:
        chave = normalizar_nome(nome)

        if chave in mapa:
            return mapa[chave]

    return None


def abrir_url_com_retentativas(
    url,
    timeout=180,
    tentativas=3
):
    ultimo_erro = None

    for tentativa in range(
        1,
        tentativas + 1
    ):
        requisicao = urllib.request.Request(
            url,
            headers={
                "User-Agent":
                "Mozilla/5.0 base-vigilancia"
            }
        )

        contexto_ssl = (
            ssl._create_unverified_context()
        )

        try:
            return urllib.request.urlopen(
                requisicao,
                timeout=timeout,
                context=contexto_ssl
            )

        except urllib.error.HTTPError as erro:
            ultimo_erro = erro

            if erro.code not in (
                408,
                429,
                500,
                502,
                503,
                504
            ):
                raise

        except (
            urllib.error.URLError,
            TimeoutError,
            OSError
        ) as erro:
            ultimo_erro = erro

        if tentativa < tentativas:
            espera = 10 * tentativa

            print(
                "Falha temporária no download:",
                repr(ultimo_erro),
                "| nova tentativa em",
                espera,
                "segundos",
                f"({tentativa + 1}/{tentativas})"
            )

            time.sleep(espera)

    raise RuntimeError(
        "Não foi possível baixar a fonte "
        f"após {tentativas} tentativas: {url}"
    ) from ultimo_erro


def baixar_csv(url, prefixo):
    print("Baixando:", url)

    temporario = Path(
        tempfile.mkstemp(
            prefix=prefixo,
            suffix=".csv"
        )[1]
    )

    with abrir_url_com_retentativas(
        url
    ) as resposta, temporario.open(
        "wb"
    ) as arquivo:

        METADADOS_FONTES[url] = {
            "data_fonte": (
                resposta.headers.get(
                    "Last-Modified"
                )
                or resposta.headers.get("Date")
                or "não informada"
            ),
            "etag": resposta.headers.get("ETag")
        }

        shutil.copyfileobj(
            resposta,
            arquivo
        )

    tamanho = temporario.stat().st_size

    print(
        "Arquivo baixado:",
        tamanho,
        "bytes"
    )

    if tamanho < 100000:
        raise RuntimeError(
            "Arquivo baixado pequeno demais. "
            "Atualização cancelada."
        )

    return temporario


def baixar_texto(url):
    with abrir_url_com_retentativas(
        url
    ) as resposta:
        return resposta.read().decode(
            "utf-8",
            errors="replace"
        )


def resolver_url_cmed():
    pagina = baixar_texto(
        PAGINA_CMED
    )

    links = re.findall(
        r'href=["\']([^"\']+\.xlsx(?:/[^"\']*)?)["\']',
        pagina,
        flags=re.IGNORECASE
    )

    candidatos = [
        urljoin(
            PAGINA_CMED,
            unescape(link)
        )
        for link in links
        if (
            "conformidade_site" in link.lower()
            and "conformidade_gov" not in link.lower()
        )
    ]

    if not candidatos:
        raise RuntimeError(
            "Link PMC - xls não encontrado "
            "na página oficial da CMED."
        )

    print(
        "Planilha CMED localizada:",
        candidatos[0]
    )

    return candidatos[0]


def baixar_xlsx(url, prefixo):
    print("Baixando:", url)

    temporario = Path(
        tempfile.mkstemp(
            prefix=prefixo,
            suffix=".xlsx"
        )[1]
    )

    with abrir_url_com_retentativas(
        url
    ) as resposta, temporario.open(
        "wb"
    ) as arquivo:

        METADADOS_FONTES[url] = {
            "data_fonte": (
                resposta.headers.get(
                    "Last-Modified"
                )
                or resposta.headers.get("Date")
                or "não informada"
            ),
            "etag": resposta.headers.get("ETag")
        }

        shutil.copyfileobj(
            resposta,
            arquivo
        )

    tamanho = temporario.stat().st_size

    print(
        "Arquivo baixado:",
        tamanho,
        "bytes"
    )

    if tamanho < 100000:
        raise RuntimeError(
            "Planilha CMED pequena demais. "
            "Atualização cancelada."
        )

    return temporario


def detectar_configuracao(arquivo):
    amostra = arquivo.read_bytes()[:200000]

    encoding = "latin-1"

    for tentativa in [
        "utf-8-sig",
        "utf-8",
        "latin-1",
        "cp1252"
    ]:
        try:
            amostra.decode(tentativa)
            encoding = tentativa
            break
        except UnicodeDecodeError:
            pass

    texto_amostra = amostra.decode(
        encoding,
        errors="replace"
    )

    delimitador = ";"

    try:
        dialeto = csv.Sniffer().sniff(
            texto_amostra,
            delimiters=";,|\t"
        )
        delimitador = dialeto.delimiter

    except csv.Error:
        pass

    print(
        "Encoding:",
        encoding,
        "| delimitador:",
        repr(delimitador)
    )

    return encoding, delimitador


def limpar_json(item):
    return {
        chave: valor
        for chave, valor in item.items()
        if valor not in ("", None)
    }


def carregar_manifesto_anterior():
    caminho = DADOS / "manifest.json"

    if not caminho.exists():
        return {}

    try:
        with caminho.open(
            "r",
            encoding="utf-8"
        ) as f:
            return json.load(f)
    except (
        OSError,
        json.JSONDecodeError
    ):
        return {}


def validar_reducao(nome, total):
    if os.environ.get(
        "PERMITIR_REDUCAO_ANORMAL"
    ) == "1":
        return

    anterior = (
        carregar_manifesto_anterior()
        .get("bases", {})
        .get(nome, {})
        .get("registros")
    )

    if not isinstance(anterior, int):
        return

    minimo = int(
        anterior * (
            1 - LIMITE_REDUCAO
        )
    )

    if total < minimo:
        percentual = (
            100 * (anterior - total)
            / anterior
        )
        raise RuntimeError(
            f"Base {nome} caiu de "
            f"{anterior} para {total} "
            f"registros ({percentual:.1f}%): "
            "atualização cancelada. "
            "Revise a fonte ou use "
            "PERMITIR_REDUCAO_ANORMAL=1 "
            "após validação manual."
        )


def metadados_base(
    url,
    total,
    fragmentos,
    prefixo,
    maior_fragmento
):
    gerado_em = datetime.now(
        timezone.utc
    ).isoformat()

    fonte = METADADOS_FONTES.get(
        url,
        {}
    )

    resultado = {
        "status": "ok",
        "fonte": url,
        "data_fonte": fonte.get(
            "data_fonte",
            "não informada"
        ),
        "gerado_em": gerado_em,
        "atualizado_em": gerado_em,
        "registros": total,
        "fragmentos": fragmentos,
        "prefixo": prefixo,
        "maior_fragmento": maior_fragmento
    }

    if fonte.get("etag"):
        resultado["etag_fonte"] = (
            fonte["etag"]
        )

    return resultado


def prefixo_processo(processo, tamanho=3):
    numero = somente_numeros(processo)

    if len(numero) >= 5 + tamanho:
        return numero[5:5 + tamanho]

    return numero[:tamanho]


def extrair_cnpj_item(item):
    cnpj = somente_numeros(
        item.get("cnpj", "")
    )

    if len(cnpj) == 14:
        return cnpj

    correspondencia = re.search(
        r"(?<!\d)(\d{14})(?!\d)",
        item.get("detentor", "")
    )

    return (
        correspondencia.group(1)
        if correspondencia
        else ""
    )


def gravar_fragmentos(destino, grupos):
    destino.mkdir(
        parents=True,
        exist_ok=True
    )

    for antigo in destino.glob("*.json"):
        antigo.unlink()

    maior = 0

    for prefixo, itens in grupos.items():

        caminho = destino / (
            prefixo + ".json"
        )

        with caminho.open(
            "w",
            encoding="utf-8"
        ) as f:

            json.dump(
                itens,
                f,
                ensure_ascii=False,
                separators=(",", ":")
            )

        maior = max(
            maior,
            caminho.stat().st_size
        )

    return maior


# ==================================================
# DISPOSITIVOS MÉDICOS
# ==================================================

def gerar_dispositivos():

    arquivo = baixar_csv(
        URL_DISPOSITIVOS,
        "anvisa_dispositivos_"
    )

    try:
        encoding, delimitador = (
            detectar_configuracao(
                arquivo
            )
        )

        grupos = defaultdict(list)

        with arquivo.open(
            "r",
            encoding=encoding,
            errors="replace",
            newline=""
        ) as f:

            leitor = csv.DictReader(
                f,
                delimiter=delimitador
            )

            if not leitor.fieldnames:
                raise RuntimeError(
                    "CSV de dispositivos "
                    "sem cabeçalho."
                )

            print(
                "Colunas dispositivos:",
                leitor.fieldnames
            )

            col_registro = achar_coluna(
                leitor.fieldnames,
                [
                    "NUMERO_REGISTRO_CADASTRO",
                    "NUMERO_REGISTRO",
                    "REGISTRO",
                    "NUM_REGISTRO",
                    "NUMERO_CADASTRO"
                ]
            )

            col_produto = achar_coluna(
                leitor.fieldnames,
                [
                    "NOME_PRODUTO",
                    "PRODUTO",
                    "NOME_COMERCIAL"
                ]
            )

            col_processo = achar_coluna(
                leitor.fieldnames,
                [
                    "NUMERO_PROCESSO",
                    "PROCESSO"
                ]
            )

            col_detentor = achar_coluna(
                leitor.fieldnames,
                [
                    "DETENTOR_REGISTRO_CADASTRO",
                    "DETENTOR_REGISTRO",
                    "DETENTOR",
                    "RAZAO_SOCIAL"
                ]
            )

            col_cnpj = achar_coluna(
                leitor.fieldnames,
                [
                    "CNPJ_DETENTOR",
                    "CNPJ",
                    "CNPJ_EMPRESA"
                ]
            )

            col_fabricante = achar_coluna(
                leitor.fieldnames,
                [
                    "NOME_FABRICANTE",
                    "FABRICANTE",
                    "FABRICANTE_LEGAL"
                ]
            )

            col_pais = achar_coluna(
                leitor.fieldnames,
                [
                    "NOME_PAIS_FABRIC",
                    "PAIS_FABRICANTE",
                    "PAIS"
                ]
            )

            col_classe = achar_coluna(
                leitor.fieldnames,
                [
                    "CLASSE_RISCO",
                    "CLASSIFICACAO_RISCO",
                    "CLASSE",
                    "CLASSIFICACAO"
                ]
            )

            col_situacao = achar_coluna(
                leitor.fieldnames,
                [
                    "SITUACAO",
                    "STATUS",
                    "SITUACAO_REGISTRO"
                ]
            )

            if (
                not col_registro
                or not col_produto
            ):
                raise RuntimeError(
                    "Colunas essenciais de "
                    "dispositivos não encontradas."
                )

            total = 0

            for linha in leitor:

                registro = somente_numeros(
                    linha.get(
                        col_registro,
                        ""
                    )
                )

                produto = texto(
                    linha.get(
                        col_produto,
                        ""
                    )
                )

                if not registro or not produto:
                    continue

                item = {
                    "registro": registro,
                    "produto": produto
                }

                if col_processo:
                    item["processo"] = (
                        somente_numeros(
                            linha.get(
                                col_processo,
                                ""
                            )
                        )
                    )

                if col_detentor:
                    item["detentor"] = texto(
                        linha.get(
                            col_detentor,
                            ""
                        )
                    )

                if col_cnpj:
                    item["cnpj"] = (
                        somente_numeros(
                            linha.get(
                                col_cnpj,
                                ""
                            )
                        )
                    )

                if col_fabricante:
                    item["fabricante"] = texto(
                        linha.get(
                            col_fabricante,
                            ""
                        )
                    )

                if col_pais:
                    item["pais"] = texto(
                        linha.get(
                            col_pais,
                            ""
                        )
                    )

                if col_classe:
                    item["classe"] = texto(
                        linha.get(
                            col_classe,
                            ""
                        )
                    )

                if col_situacao:
                    item["situacao"] = texto(
                        linha.get(
                            col_situacao,
                            ""
                        )
                    )

                item = limpar_json(item)

                prefixo = registro[:PREFIXOS_BASE["dispositivos"]]

                grupos[prefixo].append(
                    item
                )

                total += 1

        if total < 1000:
            raise RuntimeError(
                "Base de dispositivos "
                "gerou poucos registros."
            )

        validar_reducao(
            "dispositivos",
            total
        )

        maior = gravar_fragmentos(
            DADOS / "dispositivos",
            grupos
        )

        print(
            "dispositivos:", total, "registros |",
            len(grupos), "fragmentos | prefixo",
            PREFIXOS_BASE["dispositivos"], "| maior",
            f"{maior / 1024:.0f} KB"
        )

        return metadados_base(
            URL_DISPOSITIVOS,
            total,
            len(grupos),
            PREFIXOS_BASE["dispositivos"],
            maior
        )

    finally:
        arquivo.unlink(
            missing_ok=True
        )


# ==================================================
# AFE / AE
# ==================================================

def gerar_afe_ae():

    arquivo = baixar_csv(
        URL_AFE_AE,
        "anvisa_afe_ae_"
    )

    try:
        encoding, delimitador = (
            detectar_configuracao(
                arquivo
            )
        )

        grupos = defaultdict(list)

        with arquivo.open(
            "r",
            encoding=encoding,
            errors="replace",
            newline=""
        ) as f:

            leitor = csv.DictReader(
                f,
                delimiter=delimitador
            )

            if not leitor.fieldnames:
                raise RuntimeError(
                    "CSV AFE/AE sem cabeçalho."
                )

            print(
                "Colunas AFE/AE:",
                leitor.fieldnames
            )

            col_cnpj = achar_coluna(
                leitor.fieldnames,
                [
                    "CNPJ",
                    "CNPJ_EMPRESA",
                    "NU_CNPJ"
                ]
            )

            col_razao = achar_coluna(
                leitor.fieldnames,
                [
                    "RAZAO_SOCIAL",
                    "NO_RAZAO_SOCIAL",
                    "NOME_EMPRESA",
                    "EMPRESA"
                ]
            )

            col_fantasia = achar_coluna(
                leitor.fieldnames,
                [
                    "NOME_FANTASIA",
                    "NO_FANTASIA",
                    "FANTASIA"
                ]
            )

            col_autorizacao = achar_coluna(
                leitor.fieldnames,
                [
                    "NU_AUTORIZACAO",
                    "NR_AUTORIZACAO",
                    "NO_AUTORIZACAO",
                    "CO_AUTORIZACAO",
                    "NU_AUTORIZACAO_ESPECIAL",
                    "NR_AUTORIZACAO_ESPECIAL",
                    "NUMERO_AUTORIZACAO_ESPECIAL",
                    "AUTORIZACAO_ESPECIAL",
                    "NUMERO_AUTORIZACAO",
                    "NUM_AUTORIZACAO",
                    "NUMERO_AFE",
                    "NU_AFE",
                    "NUMERO_AE",
                    "NU_AE",
                    "NR_AE",
                    "AUTORIZACAO",
                    "AFE",
                    "AE"
                ]
            )

            col_processo = achar_coluna(
                leitor.fieldnames,
                [
                    "NUMERO_PROCESSO",
                    "NU_PROCESSO",
                    "PROCESSO_ANVISA",
                    "PROCESSO"
                ]
            )

            col_tipo = achar_coluna(
                leitor.fieldnames,
                [
                    "TIPO_AUTORIZACAO",
                    "TIPO_AUTORIZACAO_ESPECIAL",
                    "TIPO_AFE",
                    "TIPO_AE",
                    "TIPO"
                ]
            )

            col_classe = achar_coluna(
                leitor.fieldnames,
                [
                    "CLASSE_PRODUTO",
                    "TIPO_PRODUTO",
                    "CATEGORIA_PRODUTO",
                    "CLASSE"
                ]
            )

            col_atividade = achar_coluna(
                leitor.fieldnames,
                [
                    "ATIVIDADE",
                    "ATIVIDADES",
                    "ATIVIDADE_AUTORIZADA"
                ]
            )

            col_situacao = achar_coluna(
                leitor.fieldnames,
                [
                    "SITUACAO",
                    "STATUS"
                ]
            )

            col_data = achar_coluna(
                leitor.fieldnames,
                [
                    "DATA_AUTORIZACAO",
                    "DATA_PUBLICACAO",
                    "DATA_RESOLUCAO",
                    "DATA"
                ]
            )

            col_endereco = achar_coluna(
                leitor.fieldnames,
                [
                    "ENDERECO"
                ]
            )

            col_municipio = achar_coluna(
                leitor.fieldnames,
                [
                    "MUNICIPIO",
                    "CIDADE"
                ]
            )

            col_uf = achar_coluna(
                leitor.fieldnames,
                [
                    "UF",
                    "ESTADO"
                ]
            )

            if not col_cnpj:
                raise RuntimeError(
                    "Coluna de CNPJ da base "
                    "AFE/AE não encontrada."
                )

            total = 0

            for linha in leitor:

                cnpj = somente_numeros(
                    linha.get(
                        col_cnpj,
                        ""
                    )
                )

                if len(cnpj) != 14:
                    continue

                item = {
                    "cnpj": cnpj
                }

                if col_razao:
                    item["razao_social"] = texto(
                        linha.get(
                            col_razao,
                            ""
                        )
                    )

                if col_fantasia:
                    item["nome_fantasia"] = texto(
                        linha.get(
                            col_fantasia,
                            ""
                        )
                    )

                if col_autorizacao:
                    item["autorizacao"] = texto(
                        linha.get(
                            col_autorizacao,
                            ""
                        )
                    )

                if col_processo:
                    item["processo"] = somente_numeros(
                        linha.get(
                            col_processo,
                            ""
                        )
                    )

                if col_tipo:
                    item["tipo"] = texto(
                        linha.get(
                            col_tipo,
                            ""
                        )
                    )

                if col_classe:
                    item["classe"] = texto(
                        linha.get(
                            col_classe,
                            ""
                        )
                    )

                if col_atividade:
                    item["atividade"] = texto(
                        linha.get(
                            col_atividade,
                            ""
                        )
                    )

                if col_situacao:
                    item["situacao"] = texto(
                        linha.get(
                            col_situacao,
                            ""
                        )
                    )

                if col_data:
                    item["data"] = texto(
                        linha.get(
                            col_data,
                            ""
                        )
                    )

                if col_endereco:
                    item["endereco"] = texto(
                        linha.get(
                            col_endereco,
                            ""
                        )
                    )

                if col_municipio:
                    item["municipio"] = texto(
                        linha.get(
                            col_municipio,
                            ""
                        )
                    )

                if col_uf:
                    item["uf"] = texto(
                        linha.get(
                            col_uf,
                            ""
                        )
                    )

                item = limpar_json(item)

                prefixo = cnpj[:PREFIXOS_BASE["afe_ae"]]

                grupos[prefixo].append(
                    item
                )

                total += 1

        if total < 1000:
            raise RuntimeError(
                "Base AFE/AE gerou apenas "
                f"{total} registros."
            )

        validar_reducao(
            "afe_ae",
            total
        )

        maior = gravar_fragmentos(
            DADOS / "afe_ae",
            grupos
        )

        print(
            "afe_ae:", total, "registros |",
            len(grupos), "fragmentos | prefixo",
            PREFIXOS_BASE["afe_ae"], "| maior",
            f"{maior / 1024:.0f} KB"
        )

        return metadados_base(
            URL_AFE_AE,
            total,
            len(grupos),
            PREFIXOS_BASE["afe_ae"],
            maior
        )

    finally:
        arquivo.unlink(
            missing_ok=True
        )


# ==================================================
# MEDICAMENTOS
# ==================================================

def gerar_medicamentos():

    arquivo = baixar_csv(
        URL_MEDICAMENTOS,
        "anvisa_medicamentos_"
    )

    try:
        encoding, delimitador = (
            detectar_configuracao(
                arquivo
            )
        )

        grupos = defaultdict(list)

        with arquivo.open(
            "r",
            encoding=encoding,
            errors="replace",
            newline=""
        ) as f:

            leitor = csv.DictReader(
                f,
                delimiter=delimitador
            )

            if not leitor.fieldnames:
                raise RuntimeError(
                    "CSV de medicamentos "
                    "sem cabeçalho."
                )

            print(
                "Colunas medicamentos:",
                leitor.fieldnames
            )

            col_registro = achar_coluna(
                leitor.fieldnames,
                [
                    "NUMERO_REGISTRO_PRODUTO",
                    "NUMERO_REGISTRO",
                    "REGISTRO_PRODUTO",
                    "REGISTRO"
                ]
            )

            col_produto = achar_coluna(
                leitor.fieldnames,
                [
                    "NOME_PRODUTO",
                    "PRODUTO"
                ]
            )

            col_principio = achar_coluna(
                leitor.fieldnames,
                [
                    "PRINCIPIO_ATIVO",
                    "PRINCÍPIO_ATIVO"
                ]
            )

            col_processo = achar_coluna(
                leitor.fieldnames,
                [
                    "NUMERO_PROCESSO",
                    "PROCESSO"
                ]
            )

            col_empresa = achar_coluna(
                leitor.fieldnames,
                [
                    "EMPRESA_DETENTORA_REGISTRO",
                    "EMPRESA_DETENTORA",
                    "DETENTOR_REGISTRO",
                    "RAZAO_SOCIAL"
                ]
            )

            col_cnpj = achar_coluna(
                leitor.fieldnames,
                [
                    "CNPJ_EMPRESA",
                    "CNPJ_DETENTOR",
                    "CNPJ"
                ]
            )

            col_categoria = achar_coluna(
                leitor.fieldnames,
                [
                    "CATEGORIA_REGULATORIA",
                    "CATEGORIA_REGULATÓRIA",
                    "CATEGORIA"
                ]
            )

            col_classe = achar_coluna(
                leitor.fieldnames,
                [
                    "CLASSE_TERAPEUTICA",
                    "CLASSE_TERAPÊUTICA"
                ]
            )

            col_vencimento = achar_coluna(
                leitor.fieldnames,
                [
                    "DATA_VENCIMENTO_REGISTRO",
                    "VENCIMENTO_REGISTRO",
                    "VENCIMENTO"
                ]
            )

            col_situacao = achar_coluna(
                leitor.fieldnames,
                [
                    "SITUACAO_REGISTRO",
                    "SITUACAO",
                    "STATUS"
                ]
            )

            if (
                not col_registro
                or not col_produto
            ):
                raise RuntimeError(
                    "Não foi possível localizar "
                    "registro e produto na base "
                    "de medicamentos."
                )

            total = 0

            for linha in leitor:

                registro = somente_numeros(
                    linha.get(
                        col_registro,
                        ""
                    )
                )

                produto = texto(
                    linha.get(
                        col_produto,
                        ""
                    )
                )

                if not registro or not produto:
                    continue

                item = {
                    "registro": registro,
                    "produto": produto
                }

                if col_principio:
                    item["principio_ativo"] = texto(
                        linha.get(
                            col_principio,
                            ""
                        )
                    )

                if col_processo:
                    item["processo"] = somente_numeros(
                        linha.get(
                            col_processo,
                            ""
                        )
                    )

                if col_empresa:
                    item["detentor"] = texto(
                        linha.get(
                            col_empresa,
                            ""
                        )
                    )

                if col_cnpj:
                    item["cnpj"] = somente_numeros(
                        linha.get(
                            col_cnpj,
                            ""
                        )
                    )
                elif col_empresa:
                    item["cnpj"] = (
                        extrair_cnpj_item(item)
                    )

                if col_categoria:
                    item["categoria"] = texto(
                        linha.get(
                            col_categoria,
                            ""
                        )
                    )

                if col_classe:
                    item["classe_terapeutica"] = texto(
                        linha.get(
                            col_classe,
                            ""
                        )
                    )

                if col_vencimento:
                    item["vencimento"] = texto(
                        linha.get(
                            col_vencimento,
                            ""
                        )
                    )

                if col_situacao:
                    item["situacao"] = texto(
                        linha.get(
                            col_situacao,
                            ""
                        )
                    )

                item = limpar_json(item)

                prefixo = registro[:PREFIXOS_BASE["medicamentos"]]

                grupos[prefixo].append(
                    item
                )

                total += 1

        if total < 1000:
            raise RuntimeError(
                "Base de medicamentos gerou "
                f"apenas {total} registros."
            )

        validar_reducao(
            "medicamentos",
            total
        )

        maior = gravar_fragmentos(
            DADOS / "medicamentos",
            grupos
        )

        print(
            "medicamentos:", total, "registros |",
            len(grupos), "fragmentos | prefixo",
            PREFIXOS_BASE["medicamentos"], "| maior",
            f"{maior / 1024:.0f} KB"
        )

        return metadados_base(
            URL_MEDICAMENTOS,
            total,
            len(grupos),
            PREFIXOS_BASE["medicamentos"],
            maior
        )

    finally:
        arquivo.unlink(
            missing_ok=True
        )

# ==================================================
# CMED — IDENTIFICAÇÃO POR EAN/GTIN
# ==================================================

def texto_excel(valor):
    if valor is None:
        return ""

    if (
        isinstance(valor, float)
        and valor.is_integer()
    ):
        return str(int(valor))

    return texto(valor)


def gtin14(valor):
    numero = somente_numeros(
        texto_excel(valor)
    )

    if 8 <= len(numero) <= 14:
        return numero.zfill(14)

    return ""


def encontrar_cabecalho_cmed(planilha):
    for numero_linha, valores in enumerate(
        planilha.iter_rows(
            min_row=1,
            max_row=250,
            values_only=True
        ),
        start=1
    ):
        normalizados = {
            normalizar_nome(
                texto_excel(valor)
            )
            for valor in valores
            if valor is not None
        }

        tem_ean = (
            "EAN1" in normalizados
            or "GTIN1" in normalizados
        )

        if (
            "REGISTRO" in normalizados
            and "PRODUTO" in normalizados
            and tem_ean
        ):
            return numero_linha, [
                texto_excel(valor)
                for valor in valores
            ]

    raise RuntimeError(
        "Cabeçalho da planilha CMED "
        "não foi localizado."
    )


def gerar_cmed():
    url_cmed = resolver_url_cmed()

    arquivo = baixar_xlsx(
        url_cmed,
        "anvisa_cmed_"
    )

    livro = None

    try:
        livro = load_workbook(
            arquivo,
            read_only=True,
            data_only=True
        )

        planilha = None
        linha_cabecalho = None
        cabecalho = None

        for nome_aba in livro.sheetnames:
            candidata = livro[nome_aba]

            try:
                linha, colunas = (
                    encontrar_cabecalho_cmed(
                        candidata
                    )
                )
            except RuntimeError:
                continue

            planilha = candidata
            linha_cabecalho = linha
            cabecalho = colunas
            break

        if planilha is None:
            raise RuntimeError(
                "Nenhuma aba da CMED contém "
                "o cabeçalho esperado."
            )

        print(
            "Aba CMED:",
            planilha.title,
            "| cabeçalho na linha",
            linha_cabecalho
        )

        print(
            "Colunas CMED:",
            cabecalho
        )

        mapa = {
            normalizar_nome(nome): indice
            for indice, nome in enumerate(
                cabecalho
            )
            if nome
        }

        def indice_coluna(*possibilidades):
            for possibilidade in possibilidades:
                chave = normalizar_nome(
                    possibilidade
                )

                if chave in mapa:
                    return mapa[chave]

            return None

        col_principio = indice_coluna(
            "SUBSTÂNCIA",
            "PRINCÍPIO ATIVO"
        )
        col_cnpj = indice_coluna("CNPJ")
        col_laboratorio = indice_coluna(
            "LABORATÓRIO"
        )
        col_ggrem = indice_coluna(
            "CÓDIGO GGREM"
        )
        col_registro = indice_coluna(
            "REGISTRO"
        )
        col_ean1 = indice_coluna("EAN 1")
        col_ean2 = indice_coluna("EAN 2")
        col_ean3 = indice_coluna("EAN 3")
        col_produto = indice_coluna(
            "PRODUTO"
        )
        col_apresentacao = indice_coluna(
            "APRESENTAÇÃO"
        )
        col_classe = indice_coluna(
            "CLASSE TERAPÊUTICA"
        )
        col_tipo = indice_coluna(
            "TIPO DE PRODUTO (STATUS DO PRODUTO)",
            "TIPO DE PRODUTO"
        )
        col_tarja = indice_coluna(
            "TARJA"
        )
        col_restricao = indice_coluna(
            "RESTRIÇÃO HOSPITALAR"
        )

        essenciais = {
            "SUBSTÂNCIA": col_principio,
            "CNPJ": col_cnpj,
            "LABORATÓRIO": col_laboratorio,
            "CÓDIGO GGREM": col_ggrem,
            "REGISTRO": col_registro,
            "EAN 1": col_ean1,
            "PRODUTO": col_produto,
            "APRESENTAÇÃO": col_apresentacao,
            "CLASSE TERAPÊUTICA": col_classe,
            "TIPO DE PRODUTO": col_tipo,
            "TARJA": col_tarja,
            "RESTRIÇÃO HOSPITALAR": col_restricao
        }

        faltando = [
            nome
            for nome, encontrada in essenciais.items()
            if encontrada is None
        ]

        if faltando:
            raise RuntimeError(
                "Colunas obrigatórias da CMED "
                "não encontradas: "
                + ", ".join(faltando)
            )

        grupos = defaultdict(
            lambda: defaultdict(dict)
        )
        eans_unicos = set()
        tipos = Counter()
        tarjas = Counter()
        total = 0
        referencias_ean = 0
        restritos_hospitalares = 0

        for valores in planilha.iter_rows(
            min_row=linha_cabecalho + 1,
            values_only=True
        ):
            def obter(indice):
                if (
                    indice is None
                    or indice >= len(valores)
                ):
                    return ""

                return texto_excel(
                    valores[indice]
                )

            produto = obter(col_produto)
            registro_apresentacao = somente_numeros(
                obter(col_registro)
            )

            if (
                not produto
                or not registro_apresentacao
            ):
                continue

            eans = sorted({
                ean
                for ean in (
                    gtin14(obter(col_ean1)),
                    gtin14(obter(col_ean2)),
                    gtin14(obter(col_ean3))
                )
                if ean
            })

            if not eans:
                continue

            registro_produto = (
                registro_apresentacao[:9]
                if len(registro_apresentacao) >= 9
                else registro_apresentacao
            )

            tipo = obter(col_tipo)
            tarja = obter(col_tarja)
            restricao = obter(
                col_restricao
            )

            if "SEM TARJA" in tarja.upper():
                tarja = "Sem tarja"
            elif tarja in ("-", "- (*)"):
                tarja = "Não informada"

            codigo_restricao = (
                restricao.upper()
            )

            if codigo_restricao in (
                "S", "SIM"
            ):
                restricao = "Sim"
                restritos_hospitalares += 1
            elif codigo_restricao in (
                "N", "NAO", "NÃO"
            ):
                restricao = "Não"

            item = limpar_json({
                "eans": eans,
                "produto": produto,
                "apresentacao": obter(
                    col_apresentacao
                ),
                "principio_ativo": obter(
                    col_principio
                ),
                "registro": registro_produto,
                "registro_apresentacao": (
                    registro_apresentacao
                ),
                "laboratorio": obter(
                    col_laboratorio
                ),
                "cnpj": somente_numeros(
                    obter(col_cnpj)
                ),
                "ggrem": somente_numeros(
                    obter(col_ggrem)
                ),
                "classe_terapeutica": obter(
                    col_classe
                ),
                "tipo": tipo,
                "tarja": tarja,
                "restricao_hospitalar": restricao
            })

            chave_item = (
                item.get("ggrem", ""),
                registro_apresentacao,
                produto,
                item.get("apresentacao", "")
            )

            for ean in eans:
                grupos[
                    ean[:PREFIXOS_BASE["cmed"]]
                ][ean][chave_item] = item

                eans_unicos.add(ean)
                referencias_ean += 1

            if tipo:
                tipos[tipo] += 1

            if tarja:
                tarjas[tarja] += 1

            total += 1

        if total < 20000:
            raise RuntimeError(
                "Base CMED gerou apenas "
                f"{total} apresentações."
            )

        validar_reducao(
            "cmed",
            total
        )

        prontos = {
            prefixo: {
                ean: [
                    item
                    for _, item in sorted(
                        itens.items()
                    )
                ]
                for ean, itens in sorted(
                    mapa_eans.items()
                )
            }
            for prefixo, mapa_eans in sorted(
                grupos.items()
            )
        }

        maior = gravar_fragmentos(
            DADOS / "cmed",
            prontos
        )

        print(
            "cmed:", total, "apresentações |",
            len(eans_unicos), "EANs únicos |",
            referencias_ean, "referências |",
            len(prontos), "fragmentos | prefixo",
            PREFIXOS_BASE["cmed"], "| maior",
            f"{maior / 1024:.0f} KB"
        )

        metadados = metadados_base(
            url_cmed,
            total,
            len(prontos),
            PREFIXOS_BASE["cmed"],
            maior
        )

        metadados.update({
            "fonte_pagina": PAGINA_CMED,
            "chave": "GTIN normalizado com 14 dígitos",
            "fragmentacao": (
                "8 primeiros dígitos do GTIN-14"
            ),
            "eans_unicos": len(eans_unicos),
            "referencias_ean": referencias_ean,
            "restricao_hospitalar_sim": (
                restritos_hospitalares
            ),
            "tipos": dict(
                sorted(tipos.items())
            ),
            "tarjas": dict(
                sorted(tarjas.items())
            ),
            "campos_excluidos": [
                "preços",
                "regime de preço",
                "comercialização",
                "CAP",
                "PMVG",
                "CONFAZ",
                "ICMS"
            ]
        })

        return metadados

    finally:
        if livro is not None:
            livro.close()

        arquivo.unlink(
            missing_ok=True
        )


# ==================================================
# SANEANTES
# ==================================================

def gerar_saneantes():

    arquivo = baixar_csv(
        URL_SANEANTES,
        "anvisa_saneantes_"
    )

    try:
        encoding, delimitador = (
            detectar_configuracao(
                arquivo
            )
        )

        grupos = defaultdict(list)

        with arquivo.open(
            "r",
            encoding=encoding,
            errors="replace",
            newline=""
        ) as f:

            leitor = csv.DictReader(
                f,
                delimiter=delimitador
            )

            if not leitor.fieldnames:
                raise RuntimeError(
                    "CSV de saneantes "
                    "sem cabeçalho."
                )

            print(
                "Colunas saneantes:",
                leitor.fieldnames
            )

            col_registro = achar_coluna(
                leitor.fieldnames,
                [
                    "NU_REGISTRO_PRODUTO",
                    "NUMERO_REGISTRO",
                    "REGISTRO",
                    "NUM_REGISTRO"
                ]
            )

            col_produto = achar_coluna(
                leitor.fieldnames,
                [
                    "NO_PRODUTO",
                    "NOME_PRODUTO",
                    "PRODUTO",
                    "NOME_COMERCIAL"
                ]
            )

            col_processo = achar_coluna(
                leitor.fieldnames,
                [
                    "NU_PROCESSO",
                    "NUMERO_PROCESSO",
                    "PROCESSO"
                ]
            )

            col_empresa = achar_coluna(
                leitor.fieldnames,
                [
                    "NO_RAZAO_SOCIAL_EMPRESA",
                    "RAZAO_SOCIAL",
                    "EMPRESA",
                    "DETENTOR"
                ]
            )

            col_cnpj = achar_coluna(
                leitor.fieldnames,
                [
                    "NU_CNPJ_EMPRESA",
                    "CNPJ",
                    "CNPJ_EMPRESA"
                ]
            )

            col_categoria = achar_coluna(
                leitor.fieldnames,
                [
                    "CATEGORIA",
                    "TIPO_PRODUTO",
                    "CLASSE_PRODUTO"
                ]
            )

            col_situacao = achar_coluna(
                leitor.fieldnames,
                [
                    "ST_PRODUTO_ATIVO",
                    "SITUACAO",
                    "STATUS"
                ]
            )

            col_vencimento = achar_coluna(
                leitor.fieldnames,
                [
                    "DT_VENCIMENTO_PRODUTO",
                    "VALIDADE",
                    "DATA_VENCIMENTO",
                    "VENCIMENTO"
                ]
            )
            if (
                not col_registro
                or not col_produto
            ):
                raise RuntimeError(
                    "Não foi possível localizar "
                    "registro e produto na base "
                    "de saneantes."
                )

            total = 0

            for linha in leitor:

                registro = somente_numeros(
                    linha.get(
                        col_registro,
                        ""
                    )
                )

                produto = texto(
                    linha.get(
                        col_produto,
                        ""
                    )
                )

                if not registro or not produto:
                    continue

                item = {
                    "registro": registro,
                    "produto": produto
                }

                if col_processo:
                    item["processo"] = somente_numeros(
                        linha.get(
                            col_processo,
                            ""
                        )
                    )

                if col_empresa:
                    item["detentor"] = texto(
                        linha.get(
                            col_empresa,
                            ""
                        )
                    )

                if col_cnpj:
                    item["cnpj"] = somente_numeros(
                        linha.get(
                            col_cnpj,
                            ""
                        )
                    )

                if col_categoria:
                    item["categoria"] = texto(
                        linha.get(
                            col_categoria,
                            ""
                        )
                    )

                if col_situacao:
                    item["situacao"] = texto(
                        linha.get(
                            col_situacao,
                            ""
                        )
                    )

                if col_vencimento:
                    item["vencimento"] = texto(
                        linha.get(
                            col_vencimento,
                            ""
                        )
                    )

                item = limpar_json(item)

                prefixo = registro[:PREFIXOS_BASE["saneantes"]]

                grupos[prefixo].append(
                    item
                )

                total += 1

        if total < 1000:
            raise RuntimeError(
                "Base de saneantes gerou "
                f"apenas {total} registros."
            )

        validar_reducao(
            "saneantes",
            total
        )

        maior = gravar_fragmentos(
            DADOS / "saneantes",
            grupos
        )

        print(
            "saneantes:", total, "registros |",
            len(grupos), "fragmentos | prefixo",
            PREFIXOS_BASE["saneantes"], "| maior",
            f"{maior / 1024:.0f} KB"
        )

        return metadados_base(
            URL_SANEANTES,
            total,
            len(grupos),
            PREFIXOS_BASE["saneantes"],
            maior
        )

    finally:
        arquivo.unlink(
            missing_ok=True
        )


# ==================================================
# COSMÉTICOS
# ==================================================

def gerar_cosmeticos():

    arquivo = baixar_csv(
        URL_COSMETICOS,
        "anvisa_cosmeticos_"
    )

    try:
        encoding, delimitador = (
            detectar_configuracao(
                arquivo
            )
        )

        grupos = defaultdict(list)

        with arquivo.open(
            "r",
            encoding=encoding,
            errors="replace",
            newline=""
        ) as f:

            leitor = csv.DictReader(
                f,
                delimiter=delimitador
            )

            if not leitor.fieldnames:
                raise RuntimeError(
                    "CSV de cosméticos sem cabeçalho."
                )

            print(
                "Colunas cosméticos:",
                leitor.fieldnames
            )

            col_processo = achar_coluna(
                leitor.fieldnames,
                ["NU_PROCESSO"]
            )

            col_produto = achar_coluna(
                leitor.fieldnames,
                ["NO_PRODUTO"]
            )

            col_cnpj = achar_coluna(
                leitor.fieldnames,
                ["NU_CNPJ_EMPRESA"]
            )

            col_empresa = achar_coluna(
                leitor.fieldnames,
                ["NO_RAZAO_SOCIAL_EMPRESA"]
            )

            col_vencimento = achar_coluna(
                leitor.fieldnames,
                ["DT_VENCIMENTO"]
            )

            col_situacao = achar_coluna(
                leitor.fieldnames,
                ["ST_SITUACAO_PRODUTO"]
            )

            col_registro = achar_coluna(
                leitor.fieldnames,
                ["NU_REGISTRO"]
            )

            col_tipo = achar_coluna(
                leitor.fieldnames,
                ["DS_TIPO_PETICAO"]
            )

            col_registrado = achar_coluna(
                leitor.fieldnames,
                ["ST_REGISTRADO"]
            )

            col_atualizacao = achar_coluna(
                leitor.fieldnames,
                ["DT_ATUALIZACAO"]
            )

            essenciais = {
                "NU_PROCESSO": col_processo,
                "NO_PRODUTO": col_produto,
                "NU_CNPJ_EMPRESA": col_cnpj,
                "NO_RAZAO_SOCIAL_EMPRESA": col_empresa,
                "DT_VENCIMENTO": col_vencimento,
                "ST_SITUACAO_PRODUTO": col_situacao,
                "NU_REGISTRO": col_registro,
                "DS_TIPO_PETICAO": col_tipo,
                "ST_REGISTRADO": col_registrado,
                "DT_ATUALIZACAO": col_atualizacao
            }

            faltando = [
                nome
                for nome, coluna_encontrada
                in essenciais.items()
                if not coluna_encontrada
            ]

            if faltando:
                raise RuntimeError(
                    "Colunas obrigatórias de cosméticos "
                    "não encontradas: "
                    + ", ".join(faltando)
                )

            total = 0
            ativos = 0
            inativos = 0

            for linha in leitor:

                processo = somente_numeros(
                    linha.get(col_processo, "")
                )

                produto = texto(
                    linha.get(col_produto, "")
                )

                if not processo or not produto:
                    continue

                cnpj = somente_numeros(
                    linha.get(col_cnpj, "")
                )

                empresa = texto(
                    linha.get(col_empresa, "")
                )

                registro = somente_numeros(
                    linha.get(col_registro, "")
                )

                tipo = texto(
                    linha.get(col_tipo, "")
                )

                indicador_registrado = texto(
                    linha.get(col_registrado, "")
                )

                if not tipo:
                    tipo = (
                        "Registrado"
                        if indicador_registrado == "1"
                        else "Não informado"
                    )
                elif tipo.upper() == "REGISTRO":
                    tipo = "Registrado"
                elif tipo.upper() == "DESCARTAVEL":
                    tipo = "Descartável"

                codigo_situacao = texto(
                    linha.get(col_situacao, "")
                ).upper()

                if codigo_situacao == "S":
                    situacao = "ATIVO"
                    ativos += 1
                elif codigo_situacao == "N":
                    situacao = "INATIVO"
                    inativos += 1
                else:
                    situacao = codigo_situacao or "NÃO INFORMADA"

                vencimento = texto(
                    linha.get(col_vencimento, "")
                )
                if len(vencimento) >= 10:
                    vencimento = vencimento[:10]

                atualizado_em = texto(
                    linha.get(col_atualizacao, "")
                )
                if len(atualizado_em) >= 10:
                    atualizado_em = atualizado_em[:10]

                item = limpar_json({
                    "processo": processo,
                    "produto": produto,
                    "cnpj": cnpj,
                    "detentor": empresa,
                    "tipo": tipo,
                    "situacao": situacao,
                    "vencimento": vencimento,
                    "registro": registro,
                    "atualizado_em": atualizado_em
                })

                # O shard usa os dígitos 6 a 8 do processo. Isso evita
                # concentrar mais de um milhão de itens no prefixo 25351.
                shard = prefixo_processo(
                    processo
                )

                grupos[shard].append(
                    item
                )

                total += 1

        if total < 100000:
            raise RuntimeError(
                "Base de cosméticos gerou apenas "
                f"{total} registros."
            )

        validar_reducao(
            "cosmeticos",
            total
        )

        maior = gravar_fragmentos(
            DADOS / "cosmeticos",
            grupos
        )

        print(
            "cosmeticos:", total, "registros |",
            len(grupos), "fragmentos |",
            ativos, "ativos |",
            inativos, "inativos | maior",
            f"{maior / 1024:.0f} KB"
        )

        metadados = metadados_base(
            URL_COSMETICOS,
            total,
            len(grupos),
            PREFIXOS_BASE["cosmeticos"],
            maior
        )

        metadados.update({
            "chave": "processo",
            "fragmentacao": "dígitos 6 a 8 do processo normalizado",
            "ativos": ativos,
            "inativos": inativos
        })

        return metadados

    finally:
        arquivo.unlink(
            missing_ok=True
        )


# ==================================================
# ALIMENTOS E SUPLEMENTOS NOTIFICADOS
# ==================================================

def gerar_alimentos():

    arquivo_resultado = baixar_csv(
        URL_ALIMENTOS_RESULTADO,
        "anvisa_alimentos_resultado_"
    )

    resultados = defaultdict(list)
    total_resultados = 0

    try:
        encoding_resultado, delimitador_resultado = (
            detectar_configuracao(
                arquivo_resultado
            )
        )

        with arquivo_resultado.open(
            "r",
            encoding=encoding_resultado,
            errors="replace",
            newline=""
        ) as f:

            leitor = csv.DictReader(
                f,
                delimiter=delimitador_resultado
            )

            if not leitor.fieldnames:
                raise RuntimeError(
                    "CSV de resultados de alimentos "
                    "sem cabeçalho."
                )

            print(
                "Colunas resultados alimentos:",
                leitor.fieldnames
            )

            col_chave = achar_coluna(
                leitor.fieldnames,
                ["CO_SEQ_APRESENTACAO_PRODUTO"]
            )
            col_numero = achar_coluna(
                leitor.fieldnames,
                ["NU_APRESENTACAO_PRODUTO"]
            )
            col_registro = achar_coluna(
                leitor.fieldnames,
                ["NU_REGISTRO"]
            )
            col_validade = achar_coluna(
                leitor.fieldnames,
                ["VALIDADE"]
            )
            col_forma = achar_coluna(
                leitor.fieldnames,
                ["DS_FORMA_FISICA"]
            )
            col_situacao = achar_coluna(
                leitor.fieldnames,
                ["SITUACAO_APRESENTACAO"]
            )
            col_material = achar_coluna(
                leitor.fieldnames,
                ["MATERIAL_EMBALAGENS"]
            )
            col_embalagem = achar_coluna(
                leitor.fieldnames,
                ["TIPO_EMBALAGENS"]
            )
            col_envasadoras = achar_coluna(
                leitor.fieldnames,
                ["EMPRESAS_ENVASADORAS"]
            )
            col_internacionais = achar_coluna(
                leitor.fieldnames,
                ["EMPRESAS_INTERNACIONAIS"]
            )
            col_grupos = achar_coluna(
                leitor.fieldnames,
                ["GRUPOS_POPULACIONAIS"]
            )
            col_vias = achar_coluna(
                leitor.fieldnames,
                ["VIAS_ADMINISTRACAO"]
            )
            col_tabela = achar_coluna(
                leitor.fieldnames,
                ["TABELA_NUTRICIONAL"]
            )
            col_intolerancias = achar_coluna(
                leitor.fieldnames,
                ["INTOLERANCIAS"]
            )
            col_alergenicos = achar_coluna(
                leitor.fieldnames,
                ["ALERGENICOS"]
            )

            if not col_chave:
                raise RuntimeError(
                    "Coluna CO_SEQ_APRESENTACAO_PRODUTO "
                    "não encontrada nos resultados de alimentos."
                )

            for linha in leitor:
                chave = somente_numeros(
                    linha.get(col_chave, "")
                )

                if not chave:
                    continue

                detalhe = limpar_json({
                    "numero": texto(
                        linha.get(col_numero, "")
                    ) if col_numero else "",
                    "registro": somente_numeros(
                        linha.get(col_registro, "")
                    ) if col_registro else "",
                    "prazo_validade": texto(
                        linha.get(col_validade, "")
                    ) if col_validade else "",
                    "forma_fisica": texto(
                        linha.get(col_forma, "")
                    ) if col_forma else "",
                    "situacao": texto(
                        linha.get(col_situacao, "")
                    ) if col_situacao else "",
                    "material_embalagem": texto(
                        linha.get(col_material, "")
                    ) if col_material else "",
                    "tipo_embalagem": texto(
                        linha.get(col_embalagem, "")
                    ) if col_embalagem else "",
                    "envasadores": texto(
                        linha.get(col_envasadoras, "")
                    ) if col_envasadoras else "",
                    "empresas_internacionais": texto(
                        linha.get(col_internacionais, "")
                    ) if col_internacionais else "",
                    "grupos_populacionais": texto(
                        linha.get(col_grupos, "")
                    ) if col_grupos else "",
                    "vias_administracao": texto(
                        linha.get(col_vias, "")
                    ) if col_vias else "",
                    "tabela_nutricional": texto(
                        linha.get(col_tabela, "")
                    ) if col_tabela else "",
                    "intolerancias": texto(
                        linha.get(col_intolerancias, "")
                    ) if col_intolerancias else "",
                    "alergenicos": texto(
                        linha.get(col_alergenicos, "")
                    ) if col_alergenicos else ""
                })

                if detalhe:
                    resultados[chave].append(
                        detalhe
                    )

                total_resultados += 1

    finally:
        arquivo_resultado.unlink(
            missing_ok=True
        )

    arquivo = baixar_csv(
        URL_ALIMENTOS,
        "anvisa_alimentos_"
    )

    try:
        encoding, delimitador = (
            detectar_configuracao(
                arquivo
            )
        )

        grupos = defaultdict(list)

        with arquivo.open(
            "r",
            encoding=encoding,
            errors="replace",
            newline=""
        ) as f:

            leitor = csv.DictReader(
                f,
                delimiter=delimitador
            )

            if not leitor.fieldnames:
                raise RuntimeError(
                    "CSV de alimentos sem cabeçalho."
                )

            print(
                "Colunas alimentos:",
                leitor.fieldnames
            )

            def coluna(nome):
                return achar_coluna(
                    leitor.fieldnames,
                    [nome]
                )

            col_produto = coluna("NO_PRODUTO")
            col_regularizacao = coluna(
                "NU_REGISTRO_NOTIFICACAO_PRODUTO"
            )
            col_processo = coluna("NU_PROCESSO")
            col_empresa = coluna(
                "NO_RAZAO_SOCIAL_EMPRESA"
            )
            col_cnpj = coluna("NU_CNPJ_EMPRESA")
            col_situacao_assunto = coluna(
                "DS_SITUACAO_ASSUNTO_DOC"
            )
            col_vencimento = coluna(
                "DT_VENCIMENTO_REGISTRO"
            )
            col_regularizado_em = coluna(
                "DT_REGULARIZACAO"
            )
            col_apresentacao = coluna(
                "CO_SEQ_APRESENTACAO_PRODUTO"
            )
            col_publicacao = coluna("DT_PUBLICACAO")
            col_situacao_em = coluna("DT_SITUACAO")
            col_inicio_analise = coluna(
                "DT_INICIO_ANALISE"
            )
            col_registro_produto = coluna(
                "NU_REGISTRO_PRODUTO"
            )
            col_ativo = coluna("ST_PRODUTO_ATIVO")
            col_registro = coluna("NU_REGISTRO")
            col_marcas = coluna("MARCAS")
            col_situacao_registro = coluna(
                "SITUACAO_REGISTRO"
            )
            col_tipo = coluna("TIPO_REGULARIZACAO")
            col_categoria = coluna(
                "DS_CATEGORIA_PRODUTO"
            )
            col_alegacao = coluna(
                "DS_ALEGACAO_FUNCIONAL"
            )
            col_atualizacao = coluna("DT_CARGA_ETL")

            essenciais = {
                "NO_PRODUTO": col_produto,
                "NU_REGISTRO_NOTIFICACAO_PRODUTO": (
                    col_regularizacao
                ),
                "NU_PROCESSO": col_processo,
                "NO_RAZAO_SOCIAL_EMPRESA": col_empresa,
                "NU_CNPJ_EMPRESA": col_cnpj,
                "CO_SEQ_APRESENTACAO_PRODUTO": (
                    col_apresentacao
                ),
                "TIPO_REGULARIZACAO": col_tipo,
                "DS_CATEGORIA_PRODUTO": col_categoria,
                "DT_CARGA_ETL": col_atualizacao
            }

            faltando = [
                nome
                for nome, encontrada in essenciais.items()
                if not encontrada
            ]

            if faltando:
                raise RuntimeError(
                    "Colunas obrigatórias de alimentos "
                    "não encontradas: "
                    + ", ".join(faltando)
                )

            total = 0
            registrados = 0
            notificados = 0
            ativos = 0
            inativos = 0
            suplementos = 0
            enriquecidos = 0

            def valor(linha, coluna_encontrada):
                if not coluna_encontrada:
                    return ""
                return texto(
                    linha.get(
                        coluna_encontrada,
                        ""
                    )
                )

            def data_curta(valor_data):
                dado = texto(valor_data)

                # Registros antigos podem publicar apenas MMAAAA,
                # por exemplo 122001 para dezembro de 2001.
                if re.fullmatch(r"\d{6}", dado):
                    mes = int(dado[:2])
                    ano = int(dado[2:])

                    if 1 <= mes <= 12 and 1900 <= ano <= 2200:
                        return f"{dado[:2]}/{dado[2:]}"

                return (
                    dado[:10]
                    if len(dado) >= 10
                    else dado
                )

            for linha in leitor:
                processo = somente_numeros(
                    linha.get(col_processo, "")
                )
                produto = valor(
                    linha,
                    col_produto
                )

                if not processo or not produto:
                    continue

                regularizacao = somente_numeros(
                    linha.get(
                        col_regularizacao,
                        ""
                    )
                )

                if not regularizacao:
                    regularizacao = somente_numeros(
                        linha.get(
                            col_registro,
                            ""
                        ) if col_registro else ""
                    )

                if not regularizacao:
                    regularizacao = somente_numeros(
                        linha.get(
                            col_registro_produto,
                            ""
                        ) if col_registro_produto else ""
                    )

                tipo = valor(
                    linha,
                    col_tipo
                )

                if tipo.upper() == "NOTIFICADO":
                    tipo = "Notificado"
                    notificados += 1
                elif tipo.upper() == "REGISTRADO":
                    tipo = "Registrado"
                    registrados += 1

                codigo_ativo = valor(
                    linha,
                    col_ativo
                ).upper()

                if codigo_ativo in (
                    "S", "SIM", "1", "ATIVO"
                ):
                    ativo = True
                    ativos += 1
                elif codigo_ativo in (
                    "N", "NAO", "NÃO", "0", "INATIVO"
                ):
                    ativo = False
                    inativos += 1
                else:
                    ativo = None

                situacao_registro = valor(
                    linha,
                    col_situacao_registro
                )
                situacao_assunto = valor(
                    linha,
                    col_situacao_assunto
                )

                situacao = (
                    situacao_registro
                    or situacao_assunto
                    or (
                        "ATIVO"
                        if ativo is True
                        else "INATIVO"
                        if ativo is False
                        else "NÃO INFORMADA"
                    )
                )

                categoria = valor(
                    linha,
                    col_categoria
                )

                if "SUPLEMENT" in (
                    produto + " " + categoria
                ).upper():
                    suplementos += 1

                chave_apresentacao = somente_numeros(
                    linha.get(
                        col_apresentacao,
                        ""
                    )
                )
                apresentacoes = resultados.get(
                    chave_apresentacao,
                    []
                )

                if apresentacoes:
                    enriquecidos += 1

                item = {
                    "processo": processo,
                    "registro": regularizacao,
                    "produto": produto,
                    "marcas": valor(
                        linha,
                        col_marcas
                    ),
                    "cnpj": somente_numeros(
                        linha.get(col_cnpj, "")
                    ),
                    "detentor": valor(
                        linha,
                        col_empresa
                    ),
                    "categoria": categoria,
                    "tipo": tipo,
                    "situacao": situacao,
                    "situacao_documento": situacao_assunto,
                    "ativo": ativo,
                    "vencimento_registro": data_curta(
                        valor(linha, col_vencimento)
                    ),
                    "regularizado_em": data_curta(
                        valor(linha, col_regularizado_em)
                    ),
                    "publicado_em": data_curta(
                        valor(linha, col_publicacao)
                    ),
                    "situacao_em": data_curta(
                        valor(linha, col_situacao_em)
                    ),
                    "inicio_analise_em": data_curta(
                        valor(linha, col_inicio_analise)
                    ),
                    "alegacao_funcional": valor(
                        linha,
                        col_alegacao
                    ),
                    "atualizado_em": data_curta(
                        valor(linha, col_atualizacao)
                    )
                }

                if apresentacoes:
                    item["apresentacoes"] = (
                        apresentacoes
                    )

                if ativo is None:
                    item.pop("ativo")

                item = limpar_json(item)

                grupos[
                    prefixo_processo(
                        processo,
                        PREFIXOS_BASE["alimentos"]
                    )
                ].append(item)

                total += 1

        if total < 50000:
            raise RuntimeError(
                "Base de alimentos gerou apenas "
                f"{total} registros."
            )

        validar_reducao(
            "alimentos",
            total
        )

        maior = gravar_fragmentos(
            DADOS / "alimentos",
            grupos
        )

        print(
            "alimentos:", total, "registros |",
            len(grupos), "fragmentos |",
            registrados, "registrados |",
            notificados, "notificados |",
            suplementos, "suplementos |",
            enriquecidos, "com apresentação | maior",
            f"{maior / 1024:.0f} KB"
        )

        metadados = metadados_base(
            URL_ALIMENTOS,
            total,
            len(grupos),
            PREFIXOS_BASE["alimentos"],
            maior
        )

        fonte_resultado = METADADOS_FONTES.get(
            URL_ALIMENTOS_RESULTADO,
            {}
        )

        metadados.update({
            "fonte_resultado": URL_ALIMENTOS_RESULTADO,
            "data_fonte_resultado": fonte_resultado.get(
                "data_fonte",
                "não informada"
            ),
            "chave": "processo",
            "fragmentacao": (
                "dígitos 6 a 9 do processo normalizado"
            ),
            "registrados": registrados,
            "notificados": notificados,
            "suplementos_identificados": suplementos,
            "ativos": ativos,
            "inativos": inativos,
            "resultados_apresentacao": total_resultados,
            "produtos_enriquecidos": enriquecidos
        })

        if fonte_resultado.get("etag"):
            metadados["etag_fonte_resultado"] = (
                fonte_resultado["etag"]
            )

        return metadados

    finally:
        arquivo.unlink(
            missing_ok=True
        )


# ==================================================
# ÍNDICES AUXILIARES DE PRODUTOS
# ==================================================

def gerar_indices_produtos():
    bases = {
        "dispositivos": "registro",
        "medicamentos": "registro",
        "saneantes": "registro",
        # Alimentos são armazenados por processo; por isso a referência
        # do índice aponta para o processo, e não para o registro.
        "alimentos": "processo"
    }

    processos = defaultdict(
        lambda: defaultdict(set)
    )
    cnpjs = defaultdict(
        lambda: defaultdict(dict)
    )
    autorizacoes = defaultdict(
        lambda: defaultdict(set)
    )
    regularizacoes_alimentos = defaultdict(
        lambda: defaultdict(set)
    )

    por_base = {
        nome: {
            "processos": 0,
            "cnpjs": 0
        }
        for nome in bases
    }

    for nome, campo_referencia in bases.items():
        pasta = DADOS / nome

        for caminho in sorted(
            pasta.glob("*.json")
        ):
            with caminho.open(
                "r",
                encoding="utf-8"
            ) as f:
                itens = json.load(f)

            for item in itens:
                registro = somente_numeros(
                    item.get("registro", "")
                )

                processo = somente_numeros(
                    item.get("processo", "")
                )

                identificador = somente_numeros(
                    item.get(
                        campo_referencia,
                        ""
                    )
                )

                if not identificador:
                    continue

                referencia = (
                    nome,
                    identificador
                )

                if processo:
                    processos[
                        prefixo_processo(
                            processo
                        )
                    ][processo].add(
                        referencia
                    )
                    por_base[nome][
                        "processos"
                    ] += 1

                if (
                    nome == "alimentos"
                    and registro
                    and processo
                ):
                    regularizacoes_alimentos[
                        registro[
                            :PREFIXO_INDICE_REGULARIZACAO_ALIMENTOS
                        ]
                    ][registro].add(
                        processo
                    )

                cnpj = extrair_cnpj_item(
                    item
                )

                if len(cnpj) == 14:
                    resumo = limpar_json({
                        "b": nome,
                        "registro": registro,
                        "produto": item.get(
                            "produto",
                            ""
                        ),
                        "processo": processo,
                        "detentor": item.get(
                            "detentor",
                            ""
                        ),
                        "cnpj": cnpj,
                        "fabricante": item.get(
                            "fabricante",
                            ""
                        ),
                        "pais": item.get(
                            "pais",
                            ""
                        ),
                        "classe": item.get(
                            "classe",
                            ""
                        ),
                        "categoria": item.get(
                            "categoria",
                            ""
                        ),
                        "classe_terapeutica": (
                            item.get(
                                "classe_terapeutica",
                                ""
                            )
                        ),
                        "situacao": item.get(
                            "situacao",
                            ""
                        ),
                        "vencimento": item.get(
                            "vencimento",
                            ""
                        ) or item.get(
                            "vencimento_registro",
                            ""
                        ),
                        "tipo": item.get(
                            "tipo",
                            ""
                        ),
                        "referencia": identificador,
                        "marcas": item.get(
                            "marcas",
                            ""
                        )
                    })
                    chave_resumo = (
                        nome,
                        registro,
                        processo,
                        item.get("produto", "")
                    )
                    cnpjs[
                        cnpj[:PREFIXO_INDICE_CNPJ]
                    ][cnpj][
                        chave_resumo
                    ] = resumo
                    por_base[nome][
                        "cnpjs"
                    ] += 1

    pasta_afe = DADOS / "afe_ae"
    for caminho in sorted(
        pasta_afe.glob("*.json")
    ):
        with caminho.open(
            "r",
            encoding="utf-8"
        ) as f:
            itens = json.load(f)

        for item in itens:
            autorizacao = somente_numeros(
                item.get("autorizacao", "")
            )
            cnpj = somente_numeros(
                item.get("cnpj", "")
            )
            processo = somente_numeros(
                item.get("processo", "")
            )

            # AFE/AE usa o CNPJ como referência primária no índice de processo.
            if processo and len(cnpj) == 14:
                processos[
                    prefixo_processo(processo)
                ][processo].add(
                    ("afe_ae", cnpj)
                )

            if autorizacao and len(cnpj) == 14:
                autorizacoes[
                    autorizacao[
                        :PREFIXO_INDICE_AUTORIZACAO
                    ]
                ][autorizacao].add(cnpj)

    def preparar(grupos):
        preparados = {}

        for prefixo in sorted(grupos):
            preparados[prefixo] = {
                chave: [
                    {
                        "b": base,
                        "r": registro
                    }
                    for base, registro in sorted(
                        referencias
                    )
                ]
                for chave, referencias in sorted(
                    grupos[prefixo].items()
                )
            }

        return preparados

    processos_prontos = preparar(
        processos
    )

    cnpjs_prontos = {
        prefixo: {
            cnpj: [
                resumo
                for _, resumo in sorted(
                    resumos.items()
                )
            ]
            for cnpj, resumos in sorted(
                grupos.items()
            )
        }
        for prefixo, grupos in sorted(
            cnpjs.items()
        )
    }

    autorizacoes_prontas = {
        prefixo: {
            autorizacao: [
                {"c": cnpj}
                for cnpj in sorted(cnpjs_autorizados)
            ]
            for autorizacao, cnpjs_autorizados in sorted(
                grupos.items()
            )
        }
        for prefixo, grupos in sorted(
            autorizacoes.items()
        )
    }

    regularizacoes_alimentos_prontas = {
        prefixo: {
            regularizacao: [
                {"p": processo}
                for processo in sorted(
                    processos_regularizacao
                )
            ]
            for regularizacao, processos_regularizacao in sorted(
                grupos.items()
            )
        }
        for prefixo, grupos in sorted(
            regularizacoes_alimentos.items()
        )
    }

    gravar_fragmentos(
        DADOS / "indices" / "processos",
        processos_prontos
    )
    gravar_fragmentos(
        DADOS / "indices" / "cnpj_produtos",
        cnpjs_prontos
    )
    gravar_fragmentos(
        DADOS / "indices" / "autorizacoes",
        autorizacoes_prontas
    )
    gravar_fragmentos(
        DADOS / "indices" / "regularizacoes_alimentos",
        regularizacoes_alimentos_prontas
    )

    gerado_em = datetime.now(
        timezone.utc
    ).isoformat()

    return {
        "status": "ok",
        "gerado_em": gerado_em,
        "processos": {
            "status": "ok",
            "fragmentacao": (
                "dígitos 6 a 8 do processo "
                "normalizado"
            ),
            "chaves": sum(
                len(itens)
                for itens in processos_prontos.values()
            ),
            "referencias": sum(
                len(refs)
                for itens in processos_prontos.values()
                for refs in itens.values()
            ),
            "fragmentos": len(
                processos_prontos
            )
        },
        "cnpj_produtos": {
            "status": "ok",
            "prefixo": PREFIXO_INDICE_CNPJ,
            "fragmentacao": (
                "3 primeiros dígitos do CNPJ"
            ),
            "chaves": sum(
                len(itens)
                for itens in cnpjs_prontos.values()
            ),
            "referencias": sum(
                len(refs)
                for itens in cnpjs_prontos.values()
                for refs in itens.values()
            ),
            "fragmentos": len(
                cnpjs_prontos
            )
        },
        "autorizacoes": {
            "status": (
                "ok"
                if autorizacoes_prontas
                else "indisponivel"
            ),
            "prefixo": PREFIXO_INDICE_AUTORIZACAO,
            "observacao": (
                "Índice por número de AFE/AE."
                if autorizacoes_prontas
                else "A fonte aberta não publicou número de autorização reconhecível."
            ),
            "chaves": sum(
                len(itens)
                for itens in autorizacoes_prontas.values()
            ),
            "referencias": sum(
                len(refs)
                for itens in autorizacoes_prontas.values()
                for refs in itens.values()
            ),
            "fragmentos": len(
                autorizacoes_prontas
            )
        },
        "regularizacoes_alimentos": {
            "status": (
                "ok"
                if regularizacoes_alimentos_prontas
                else "indisponivel"
            ),
            "prefixo": (
                PREFIXO_INDICE_REGULARIZACAO_ALIMENTOS
            ),
            "fragmentacao": (
                "3 primeiros dígitos do número de "
                "registro ou notificação"
            ),
            "chaves": sum(
                len(itens)
                for itens in regularizacoes_alimentos_prontas.values()
            ),
            "referencias": sum(
                len(refs)
                for itens in regularizacoes_alimentos_prontas.values()
                for refs in itens.values()
            ),
            "fragmentos": len(
                regularizacoes_alimentos_prontas
            )
        },
        "por_base": por_base
    }


# ==================================================
# MANIFEST
# ==================================================

def gerar_manifesto(
    dispositivos,
    afe_ae,
    medicamentos,
    cmed,
    saneantes,
    cosmeticos,
    alimentos,
    indices
):

    DADOS.mkdir(
        parents=True,
        exist_ok=True
    )

    caminho_manifesto = DADOS / "manifest.json"
    manifesto_anterior = {}

    if caminho_manifesto.exists():
        try:
            manifesto_anterior = json.loads(
                caminho_manifesto.read_text(
                    encoding="utf-8"
                )
            )
        except (OSError, json.JSONDecodeError) as erro:
            raise RuntimeError(
                "Manifesto principal anterior inválido."
            ) from erro

    bases_principais = {
        "dispositivos",
        "afe_ae",
        "medicamentos",
        "cmed",
        "saneantes",
        "cosmeticos",
        "alimentos"
    }
    indices_principais = {
        "status",
        "gerado_em",
        "processos",
        "cnpj_produtos",
        "autorizacoes",
        "regularizacoes_alimentos",
        "por_base"
    }

    bases_preservadas = {
        nome: dados
        for nome, dados in (
            manifesto_anterior.get("bases", {})
            if isinstance(manifesto_anterior, dict)
            else {}
        ).items()
        if nome not in bases_principais
    }
    indices_preservados = {
        nome: dados
        for nome, dados in (
            manifesto_anterior.get("indices", {})
            if isinstance(manifesto_anterior, dict)
            else {}
        ).items()
        if nome not in indices_principais
    }

    manifesto = {
        "versao_esquema": 2,

        "projeto":
        "Base Vigilância Sanitária",

        "gerado_em":
        datetime.now(
            timezone.utc
        ).isoformat(),

        "bases": {
            "dispositivos":
            dispositivos,

            "afe_ae":
            afe_ae,

            "medicamentos":
             medicamentos,

            "cmed":
            cmed,

            "saneantes":
            saneantes,

            "cosmeticos":
            cosmeticos,

            "alimentos":
            alimentos
        },

        "indices": indices
    }

    manifesto["bases"].update(
        bases_preservadas
    )
    manifesto["indices"].update(
        indices_preservados
    )

    with (
        DADOS / "manifest.json"
    ).open(
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            manifesto,
            f,
            ensure_ascii=False,
            indent=2
        )


# ==================================================
# EXECUÇÃO
# ==================================================

if __name__ == "__main__":

    print(
        "Versão do gerador:",
        VERSAO_GERADOR
    )

    print(
        "=== DISPOSITIVOS ==="
    )

    dispositivos = gerar_dispositivos()

    print(
        "=== AFE / AE ==="
    )

    afe_ae = gerar_afe_ae()

    print(
        "=== MEDICAMENTOS ==="
    )

    medicamentos = gerar_medicamentos()

    print(
        "=== CMED — EAN/GTIN ==="
    )

    cmed = gerar_cmed()

    print(
    "=== SANEANTES ==="
    )

    saneantes = gerar_saneantes()

    print(
        "=== COSMÉTICOS ==="
    )

    cosmeticos = gerar_cosmeticos()

    print(
        "=== ALIMENTOS E SUPLEMENTOS ==="
    )

    alimentos = gerar_alimentos()

    print(
        "=== ÍNDICES AUXILIARES ==="
    )

    indices = gerar_indices_produtos()

    gerar_manifesto(
        dispositivos,
        afe_ae,
        medicamentos,
        cmed,
        saneantes,
        cosmeticos,
        alimentos,
        indices
    )

    print(
        "Todas as bases foram "
        "geradas com sucesso."
    )
