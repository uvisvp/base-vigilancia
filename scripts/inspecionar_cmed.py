from collections import Counter
from html import unescape
from pathlib import Path
from urllib.parse import urljoin
import re
import shutil
import ssl
import tempfile
import urllib.request

from openpyxl import load_workbook


PAGINA_CMED = (
    "https://www.gov.br/anvisa/pt-br/assuntos/"
    "medicamentos/cmed/precos"
)

EAN_TESTE = "07896641816864"


CANDIDATOS = {
    "principio_ativo": [
        "PRINCÍPIO ATIVO", "PRINCIPIO ATIVO", "SUBSTÂNCIA", "SUBSTANCIA",
    ],
    "cnpj": ["CNPJ"],
    "laboratorio": ["LABORATÓRIO", "LABORATORIO"],
    "ggrem": ["CÓDIGO GGREM", "CODIGO GGREM", "GGREM"],
    "registro": ["REGISTRO"],
    "ean1": ["EAN 1", "EAN1", "GTIN 1", "GTIN1"],
    "ean2": ["EAN 2", "EAN2", "GTIN 2", "GTIN2"],
    "ean3": ["EAN 3", "EAN3", "GTIN 3", "GTIN3"],
    "produto": ["PRODUTO"],
    "apresentacao": ["APRESENTAÇÃO", "APRESENTACAO"],
    "classe_terapeutica": [
        "CLASSE TERAPÊUTICA", "CLASSE TERAPEUTICA",
    ],
    "tipo": ["TIPO DE PRODUTO", "TIPO PRODUTO"],
    "tarja": ["TARJA"],
    "restricao_hospitalar": [
        "RESTRIÇÃO HOSPITALAR", "RESTRICAO HOSPITALAR",
    ],
}


def normalizar_nome(valor):
    texto = str(valor or "").upper()
    trocas = str.maketrans(
        "ÁÀÂÃÉÈÊÍÌÎÓÒÔÕÚÙÛÇ",
        "AAAAEEEIIIOOOOUUUC",
    )
    texto = texto.translate(trocas)
    return re.sub(r"[^A-Z0-9]", "", texto)


def texto_excel(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return re.sub(r"\s+", " ", str(valor).strip())


def somente_numeros(valor):
    return re.sub(r"\D", "", texto_excel(valor))


def gtin_canonico(valor):
    numero = somente_numeros(valor)
    if 8 <= len(numero) <= 14:
        return numero.zfill(14)
    return numero


def gtin_valido(valor):
    numero = somente_numeros(valor)
    if len(numero) not in (8, 12, 13, 14):
        return False

    corpo = numero[:-1]
    verificador = int(numero[-1])
    soma = 0
    peso = 3

    for digito in reversed(corpo):
        soma += int(digito) * peso
        peso = 1 if peso == 3 else 3

    calculado = (10 - soma % 10) % 10
    return calculado == verificador


def baixar_texto(url):
    requisicao = urllib.request.Request(
        url,
        headers={"User-Agent": "Base-Vigilancia/1.0"},
    )
    contexto = ssl._create_unverified_context()

    with urllib.request.urlopen(
        requisicao,
        timeout=180,
        context=contexto,
    ) as resposta:
        return resposta.read().decode("utf-8", errors="replace")


def resolver_url_planilha():
    pagina = baixar_texto(PAGINA_CMED)
    links = re.findall(
        r'href=["\']([^"\']+\.xlsx(?:/[^"\']*)?)["\']',
        pagina,
        flags=re.IGNORECASE,
    )

    links = [
        urljoin(PAGINA_CMED, unescape(link))
        for link in links
    ]

    candidatos = [
        link
        for link in links
        if "conformidade_site" in link.lower()
        and "conformidade_gov" not in link.lower()
    ]

    if not candidatos:
        raise RuntimeError(
            "Não foi localizado o link PMC - xls na página da CMED."
        )

    return candidatos[0]


def baixar_planilha(url):
    temporario = tempfile.NamedTemporaryFile(
        prefix="cmed_",
        suffix=".xlsx",
        delete=False,
    )
    caminho = Path(temporario.name)
    temporario.close()

    print("Baixando planilha CMED:", url, flush=True)

    requisicao = urllib.request.Request(
        url,
        headers={"User-Agent": "Base-Vigilancia/1.0"},
    )
    contexto = ssl._create_unverified_context()

    with urllib.request.urlopen(
        requisicao,
        timeout=600,
        context=contexto,
    ) as resposta, caminho.open("wb") as destino:
        shutil.copyfileobj(resposta, destino)

    tamanho = caminho.stat().st_size
    print("Arquivo baixado:", tamanho, "bytes", flush=True)

    if tamanho < 100_000:
        raise RuntimeError(
            f"Planilha CMED parece incompleta: {tamanho} bytes."
        )

    return caminho


def encontrar_cabecalho(planilha):
    for numero_linha, valores in enumerate(
        planilha.iter_rows(
            min_row=1,
            max_row=40,
            values_only=True,
        ),
        start=1,
    ):
        normalizados = {
            normalizar_nome(valor)
            for valor in valores
            if valor is not None
        }

        if {
            "REGISTRO", "PRODUTO", "APRESENTACAO"
        }.issubset(normalizados) and (
            "EAN1" in normalizados or "GTIN1" in normalizados
        ):
            return numero_linha, [
                texto_excel(valor)
                for valor in valores
            ]

    raise RuntimeError(
        "Cabeçalho da planilha CMED não foi localizado."
    )


def mapear_colunas(cabecalho):
    mapa = {
        normalizar_nome(nome): indice
        for indice, nome in enumerate(cabecalho)
        if nome
    }

    resultado = {}

    for campo, candidatos in CANDIDATOS.items():
        resultado[campo] = None

        for candidato in candidatos:
            chave = normalizar_nome(candidato)
            if chave in mapa:
                resultado[campo] = mapa[chave]
                break

    return resultado


def inspecionar(caminho):
    livro = load_workbook(
        caminho,
        read_only=True,
        data_only=True,
    )

    print("ABAS:", livro.sheetnames, flush=True)

    planilha = None
    linha_cabecalho = None
    cabecalho = None

    for nome_aba in livro.sheetnames:
        candidata = livro[nome_aba]
        try:
            linha, colunas = encontrar_cabecalho(candidata)
        except RuntimeError:
            continue

        planilha = candidata
        linha_cabecalho = linha
        cabecalho = colunas
        break

    if planilha is None:
        raise RuntimeError(
            "Nenhuma aba contém o cabeçalho esperado da CMED."
        )

    print("ABA UTILIZADA:", planilha.title, flush=True)
    print("LINHA DO CABEÇALHO:", linha_cabecalho, flush=True)
    print("COLUNAS CMED:", cabecalho, flush=True)

    mapeamento = mapear_colunas(cabecalho)
    print("MAPEAMENTO PROPOSTO:", mapeamento, flush=True)

    essenciais = (
        "registro", "produto", "apresentacao", "laboratorio", "ean1"
    )
    faltando = [
        campo for campo in essenciais
        if mapeamento.get(campo) is None
    ]

    if faltando:
        raise RuntimeError(
            "Campos essenciais ausentes: " + ", ".join(faltando)
        )

    total = 0
    com_ean = 0
    eans_validos = 0
    eans = set()
    tipos = Counter()
    tarjas = Counter()
    restricoes = Counter()
    amostras = []
    encontrados_teste = []

    for valores in planilha.iter_rows(
        min_row=linha_cabecalho + 1,
        values_only=True,
    ):
        def obter(campo):
            indice = mapeamento.get(campo)
            if indice is None or indice >= len(valores):
                return ""
            return texto_excel(valores[indice])

        produto = obter("produto")
        registro = somente_numeros(obter("registro"))

        if not produto or not registro:
            continue

        item = {
            "principio_ativo": obter("principio_ativo"),
            "cnpj": somente_numeros(obter("cnpj")),
            "laboratorio": obter("laboratorio"),
            "ggrem": somente_numeros(obter("ggrem")),
            "registro": registro,
            "ean1": somente_numeros(obter("ean1")),
            "ean2": somente_numeros(obter("ean2")),
            "ean3": somente_numeros(obter("ean3")),
            "produto": produto,
            "apresentacao": obter("apresentacao"),
            "classe_terapeutica": obter("classe_terapeutica"),
            "tipo": obter("tipo"),
            "tarja": obter("tarja"),
            "restricao_hospitalar": obter("restricao_hospitalar"),
        }

        item = {
            chave: valor
            for chave, valor in item.items()
            if valor
        }

        total += 1
        eans_item = [
            item.get(campo, "")
            for campo in ("ean1", "ean2", "ean3")
            if item.get(campo)
        ]

        if eans_item:
            com_ean += 1

        for ean in eans_item:
            eans.add(gtin_canonico(ean))
            if gtin_valido(ean):
                eans_validos += 1

            if gtin_canonico(ean) == gtin_canonico(EAN_TESTE):
                encontrados_teste.append(item)

        if item.get("tipo"):
            tipos[item["tipo"]] += 1
        if item.get("tarja"):
            tarjas[item["tarja"]] += 1
        if item.get("restricao_hospitalar"):
            restricoes[item["restricao_hospitalar"]] += 1

        if len(amostras) < 5:
            amostras.append(item)

    livro.close()

    print("TOTAL DE APRESENTAÇÕES:", total, flush=True)
    print("APRESENTAÇÕES COM EAN:", com_ean, flush=True)
    print("EANS ÚNICOS:", len(eans), flush=True)
    print("EANS COM DÍGITO VERIFICADOR VÁLIDO:", eans_validos, flush=True)
    print("TIPOS (30 mais frequentes):", tipos.most_common(30), flush=True)
    print("TARJAS:", tarjas.most_common(30), flush=True)
    print("RESTRIÇÃO HOSPITALAR:", restricoes.most_common(30), flush=True)
    print("PRIMEIRAS LINHAS MAPEADAS:", amostras, flush=True)
    print(
        f"RESULTADO DO EAN DE TESTE {EAN_TESTE}:",
        encontrados_teste,
        flush=True,
    )


def main():
    url = resolver_url_planilha()
    print("URL PMC XLS LOCALIZADA:", url, flush=True)
    caminho = baixar_planilha(url)

    try:
        inspecionar(caminho)
    finally:
        caminho.unlink(missing_ok=True)

    print("Inspeção CMED concluída com sucesso.", flush=True)


if __name__ == "__main__":
    main()
